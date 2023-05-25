import requests
import icalendar
import re
import datetime
import win32com.client
import openpyxl
import os
from Constant import *

#是否写入本地html
IsLocalHtml = False

#用于登录的URL以及UA
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36 Edg/114.0.0.0'
LOGIN_URL = 'https://jwgl.bupt.edu.cn/jsxsd/'
GET_API = 'https://jwgl.bupt.edu.cn/jsxsd/xskb/xskb_list.do'
POST_URL = 'https://jwgl.bupt.edu.cn/jsxsd/xk/LoginToXk'
KBJCMSID = { #由[https://jwgl.bupt.edu.cn/jsxsd/framework/xsMain.jsp]的html代码得到的'课表节次模式id'
    '0': '9475847A3F3033D1E05377B5030AA94D',  #默认节次模式
    '1': '857D7D63586B4AF4B4C59B1DFEEB56E4'  #海南校区课表时间
}
KBLB = { #由[https://jwgl.bupt.edu.cn/jsxsd/framework/xsMain.jsp]的html代码得到的'课表类别'
    '0': 'kckb',  #课程课表
    '1': 'tzdkb'  #通知单课表
}

def GetXNXQ(): #获取学年学期，形如'2022-2023-2'，2-7月为第二学期，8-次年1月为第一学期
    CurrentYear = datetime.datetime.now().year
    CurrentMonth = datetime.datetime.now().month
    if CurrentMonth >= 2 and CurrentMonth <= 7:
        XNXQ = str(CurrentYear - 1) + '-' + str(CurrentYear) + '-2'
    else:
        XNXQ = str(CurrentYear) + '-' + str(CurrentYear + 1) + '-1'
    return XNXQ

def ProcessCK(CK):  #处理cookies，将cookies转换为字符串
    Cookie = ''
    for Name, Value in CK:
        Cookie += f'{Name}={Value};'
    return Cookie

#以下加密算法参考[https://github.com/LAWTED/BUPT-Auto-Syllabu/blob/main/process.py]得到
KEYSTRING = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/="
def EncodeInp(Input):
    Output = ''
    Chr1, Chr2, Chr3 = '', '', ''
    Enc1, Enc2, Enc3 = '', '', ''
    i = 0
    while True:
        Chr1 = ord(Input[i])
        i += 1
        Chr2 = ord(Input[i]) if i < len(Input) else 0
        i += 1
        Chr3 = ord(Input[i]) if i < len(Input) else 0
        i += 1
        Enc1 = Chr1 >> 2
        Enc2 = ((Chr1 & 3) << 4) | (Chr2 >> 4)
        Enc3 = ((Chr2 & 15) << 2) | (Chr3 >> 6)
        Enc4 = Chr3 & 63
        if Chr2 == 0:
            Enc3 = Enc4 = 64
        elif Chr3 == 0:
            Enc4 = 64
        Output = Output + KEYSTRING[Enc1] + KEYSTRING[Enc2] + KEYSTRING[Enc3] + KEYSTRING[Enc4]
        Chr1 = Chr2 = Chr3 = ''
        Enc1 = Enc2 = Enc3 = Enc4 = ''
        if i >= len(Input): break
    return Output

Encoded = EncodeInp(SchoolID) + "%%%" + EncodeInp(Password_jwgl) #加密后的学号和密码
Xueqi = GetXNXQ() #获取学年学期，形如'2022-2023-2'
UserData = { #用户数据，保存在Constant.py中，根据需要修改
    'userAccount': SchoolID,
    'userPassword': Password_jwgl,
    'encoded': Encoded
}

#登录教务系统
print('登录北邮教务管理系统')
Session = requests.Session()
Login = Session.get(url=LOGIN_URL, headers={'User-Agent': USER_AGENT})
Cookies1 = Login.cookies.items() #处理cookies
Cookie = ProcessCK(Cookies1)

#设置headers，先post用户数据
Headers_UserData = {
    'Host': 'jwgl.bupt.edu.cn',
    'Referer': 'https://jwgl.bupt.edu.cn/jsxsd/xk/LoginToXk?method=exit&tktime=1631723647000',
    'User-Agent': USER_AGENT,
    'cookie': Cookie
}
Session.post(url=POST_URL, data=UserData, headers=Headers_UserData) #由LOGIN_URL的html决定是post方法，action='/jsxsd/xk/LoginToXk'
Cookies2 = Login.cookies.items() #登录后cookie可能变了，重新处理一下
Cookie_Logged_In = ProcessCK(Cookies2)

#设置headers
Headers = {'User-Agent': USER_AGENT, 'cookie': Cookie_Logged_In}
#周次（默认全部），学年学期（默认当前学期），课表节次id，课表类别
Info = {
    'zc': '',
    'xnxq01id': Xueqi,
    'kbjcmsid': KBJCMSID['0'],
    'kblb': KBLB['0']
}
Url = 'https://jwgl.bupt.edu.cn/jsxsd/xskb/xskb_list.do'
Data = requests.post(url=Url, headers=Headers, data=Info)  #获取课表html，检查是否成功
if "请先登录系统" in Data.text:
    print("登录失败，请检查学号和密码以及网络连接，并重新运行程序")
    exit()
else:
    print("登录成功，正在处理")
    print("已获取到课表信息")
xlsUrl = f'https://jwgl.bupt.edu.cn/jsxsd/xskb/xskb_print.do?xnxq01id={Xueqi}&zc=&kbjcmsid=9475847A3F3033D1E05377B5030AA94D' #获取课表xls文件的url
Excel = requests.post(url=xlsUrl, headers=Headers) #获取课表xls文件
f = open(f'学生个人课表_{SchoolID}.xls', 'wb')
f.write(Excel.content)  #写入本地xls文件
f.close()

#可选，将.do写入本地html文件
if IsLocalHtml:
    print("正在写入本地html")
    try:
        with open(f'./{datetime.datetime.now().strftime("%Y-%m-%d-%H-%M")}.html', 'w', encoding='utf-8') as file:
            file.write(Data.text)
            print(f'[success]')
    except Exception:
        print("生成文件失败")

#将课表xls文件转换为xlsx文件
Fname = f'{os.getcwd()}/学生个人课表_{SchoolID}.xls'
excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(Fname)
wb.SaveAs(Fname + 'x', FileFormat=51)  #FileFormat = 51 is for .xlsx extension
wb.Close()  #FileFormat = 56 is for .xls extension
excel.Application.Quit()

#-----------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------
#以下Excel处理代码来自https://github.com/Guest-Liang/BUPT-iOSCalendar-Excel
#定义课程开始时间
StartTime = [
    datetime.time(8, 0, 0),   datetime.time(8, 50, 0),  datetime.time(9, 50, 0), 
    datetime.time(10, 40, 0), datetime.time(11, 30, 0), datetime.time(13, 00, 0), 
    datetime.time(13, 50, 0), datetime.time(14, 45, 0), datetime.time(15, 40, 0), 
    datetime.time(16, 35, 0), datetime.time(17, 25, 0), datetime.time(18, 30, 0), 
    datetime.time(19, 20, 0), datetime.time(20, 10, 0)
]
#定义课程结束时间
EndTime = [
    datetime.time(8, 45, 0),  datetime.time(9, 35, 0),  datetime.time(10, 35, 0), 
    datetime.time(11, 25, 0), datetime.time(12, 15, 0), datetime.time(13, 45, 0), 
    datetime.time(14, 35, 0), datetime.time(15, 30, 0), datetime.time(16, 25, 0), 
    datetime.time(17, 20, 0), datetime.time(18, 10, 0), datetime.time(19, 15, 0), 
    datetime.time(20, 5, 0),  datetime.time(20, 55, 0)
]

#找到字符串中某个关键值的所有索引，存放在list(int)中
def GetElementIndex(char, string):
    return [idx.start() for idx in re.finditer(char, string)]

#将上课周数转为list(int)
def ChangeIntoList_int(s):
    ranges = re.findall(r'(\d+)-(\d+)', s)
    for start, end in ranges:
        s = s.replace(f'{start}-{end}', ','.join(map(str, range(int(start), int(end) + 1))))
    return list(map(int, s.split(',')))

#获取学号，打开xlsx文件
userid = SchoolID
WorkBook = openpyxl.load_workbook(filename=f"./学生个人课表_{userid}.xlsx")
Sheet = WorkBook.active

print("-------------------------")
print("您的信息为：")
print("课程表所属人：", end="")
StudentName = Sheet['A1'].value.replace("北京邮电大学 ", "").replace(" 学生个人课表", "")
print(StudentName)
print("学年：", end="")
SchoolYear = Sheet['A2'].value[5:16]
print(SchoolYear)

#输入学期的第一周的周一日期
#StartDate=datetime.date(2023, 2, 20)
# StartDate = datetime.datetime.strptime(input("输入学期的第一周的周一的日期，以YYYY-MM-DD格式\n"), '%Y-%m-%d').date()
# while StartDate.isoweekday() != 1:
#     StartDate = datetime.datetime.strptime(input("日期并非周一！请以YYYY-MM-DD格式输入\n"), '%Y-%m-%d').date()
StartDate=datetime.datetime.strptime(f'{StartDate}', '%Y-%m-%d').date()
print("正在处理")

#制作部分
MyCalendar = icalendar.Calendar()
MyCalendar.add('PRODID', '-//MY_CALENDAR_PRODUCT//GL//')
MyCalendar.add('VERSION', '2.0') #固定属性，版本2.0
MyCalendar.add('CALSCALE', 'GREGORIAN')  #公历
MyCalendar.add('METHOD', 'PUBLISH')
MyCalendar.add('X-WR-CALNAME', f'{SchoolYear}') #通用属性，日历名称，默认为学年
MyCalendar.add('X-WR-TIMEZONE', 'Asia/Shanghai') #通用属性，指定时区
MyCalendar.add('X-APPLE-CALENDAR-COLOR', '#E1FFFF') #私有属性，指定Apple日历颜色，可自己更改，填入十六进制代码
for Column in range(2, 9):
    for Row in range(4, 18):
        CellBR = GetElementIndex("\n", Sheet.cell(row=Row, column=Column).value) #根据换行符位置进行拆分
        for i in range(int(len(CellBR) / 5)): #拆分 课程、教师名字、上课周数、上课教室、上课节次
            Course = Sheet.cell(Row, Column).value[CellBR[5 * i] + 1:CellBR[5 * i + 1]]
            TeacherName = Sheet.cell(Row, Column).value[CellBR[5 * i + 1] + 1:CellBR[5 * i + 2]]
            ClassWeeks = Sheet.cell(Row, Column).value[CellBR[5 * i + 2] + 1:CellBR[5 * i + 3]]
            Classroom = Sheet.cell(Row, Column).value[CellBR[5 * i + 3] + 1:CellBR[5 * i + 4]]
            if i == int(len(CellBR) / 5) - 1:
                LessonNum = Sheet.cell(Row, Column).value[CellBR[5 * i + 4] + 1:]
            else:
                LessonNum = Sheet.cell(Row, Column).value[CellBR[5 * i + 4] + 1:CellBR[5 * i + 5]]
            ListLessonNum = LessonNum.replace("[", "").replace("]", "").replace("节", "").split("-")
            ListLessonNum = list(map(int, ListLessonNum))
            if (Row - 3 == ListLessonNum[0]): #是第一节课才添加，下一节跳过（意思是默认合并课程，避免导致相邻、连着的课程重复提醒）
                ListClassWeeks = ChangeIntoList_int(ClassWeeks.replace("[周]", ""))
                for j in range(len(ListClassWeeks)):
                    MyEvent = icalendar.Event()
                    MyEvent.add('SUMMARY', Course + ' ' + Classroom) #事件名称：课程名加教室
                    MyEvent.add('DTSTAMP', datetime.datetime.today())
                    MyEvent.add('DTSTART', datetime.datetime.combine(StartDate + datetime.timedelta(weeks=ListClassWeeks[j] - 1, days=Column - 2), StartTime[ListLessonNum[0] - 1]))
                    MyEvent.add('DTEND', datetime.datetime.combine(StartDate + datetime.timedelta(weeks=ListClassWeeks[j] - 1, days=Column - 2), EndTime[ListLessonNum[-1] - 1]))
                    MyEvent.add('DESCRIPTION', TeacherName) #教师姓名写在备注里
                    MyAlarm = icalendar.Alarm() #添加提醒作为事件的附加属性
                    MyAlarm.add('TRIGGER', datetime.timedelta(minutes=-10)) #提前10分钟提醒
                    MyAlarm.add('ACTION', "DISPLAY") #通知提醒
                    MyAlarm.add('DESCRIPTION', Course) #提醒内容：课程名称
                    MyEvent.add_component(MyAlarm)
                    MyCalendar.add_component(MyEvent)
                    del MyAlarm, MyEvent
                del TeacherName, ClassWeeks, Classroom, LessonNum, Course, ListClassWeeks
        del CellBR
try:
    with open(f'TimeTable_{StudentName}.ics', 'wb') as file:
        file.write(MyCalendar.to_ical())
        print('[Success]')
        del MyCalendar
except Exception:
    print("生成文件失败，请重试")